# coding=utf-8

"""
@fileName       :   columns.py
@data           :   2024/11/15
@author         :   jiangmenggui@hosonsoft.com
"""
import datetime
import re
from typing import Callable, Any, Literal

from dateutil import parser
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class ExcelColumn:

    def __init__(
            self, name: str, /, *, required=False, default=None, default_factory: Callable[[], Any] | None = None, **kwargs
    ):
        self.name = name
        self.required = required
        self.default_factory = default_factory or (lambda: default)
        self.kwargs = kwargs

    def validate(self, value):
        pass

    def get_data_validation(self, columnIndex):
        return None


class StringColumn(ExcelColumn):

    def __init__(
            self, name: str, /, *,
            required=False, default=None, default_factory=None, min_length=None, max_length=None, **kwargs
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory,
            min_length=min_length, max_length=max_length, **kwargs
        )

    def validate(self, value):
        if value is None:
            value = self.default_factory()
        if self.required and value is None:
            raise ValueError(f'字段 {self.name} 不能为空')
        if value is None:
            return value
        min_length = self.kwargs.get('min_length')
        max_length = self.kwargs.get('max_length')
        if min_length is not None and len(value) < min_length:
            raise ValueError(f'字段 {self.name} 长度不能小于 {min_length}')
        if max_length is not None and len(value) > max_length:
            raise ValueError(f'字段 {self.name} 长度不能大于 {max_length}')
        return str(value)

    def get_data_validation(self, columnIndex):
        operator: Literal[
            "between", "notBetween", "equal", "notEqual",
            "lessThan", "lessThanOrEqual", "greaterThan",  "greaterThanOrEqual"
        ]
        if self.kwargs.get('max_length') is not None and self.kwargs.get('min_length') is not None:
            operator = 'between'
            formula = dict(formula1=self.kwargs.get('min_length'), formula2=self.kwargs.get('max_length'),)
            errorMessage = f'{self.name}长度必须介于{self.kwargs.get("min_length")}~{self.kwargs.get("max_length")}之间'
        elif self.kwargs.get('min_length') is not None and self.kwargs.get('min_length', 0) > 0:
            operator = 'greaterThanOrEqual'
            formula = {"formula1": self.kwargs.get('min_length')}
            errorMessage = f"{self.name}长度不能少于{self.kwargs.get('min_length')}位"
        elif self.kwargs.get('max_length') is not None and self.kwargs.get('max_length', 0) > 0:
            operator = 'lessThanOrEqual'
            formula = {"formula1": self.kwargs.get('max_length')}
            errorMessage = f"{self.name}长度不能超过{self.kwargs.get('max_length')}位"
        else:
            return None
        dv = DataValidation(
            type='textLength',
            operator=operator,
            **formula,
            showErrorMessage=True,
            showInputMessage=False
        )
        dv.error = errorMessage
        dv.errorTitle = "无效的输入"
        columnName = get_column_letter(columnIndex)
        dv.add(f'{columnName}{2}:{columnName}{10000}')  # 数据验证区域,2-11行
        return dv


class RegexColumn(StringColumn):

    def __init__(
            self, name: str, pattern: str, /, *,
            required=False, default=None, default_factory=None, min_length=None, max_length=None
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, min_length=min_length,
            max_length=max_length, pattern=pattern
        )

    def validate(self, value):
        value = super().validate(value)
        if value is None:
            return value
        if not re.match(self.kwargs['pattern'], value):
            raise ValueError(f'字段 {self.name} 格式不正确')
        return value


class EmailColumn(RegexColumn):

    def __init__(
            self, name: str, /, *,
            required=False, default=None, default_factory=None, min_length=None, max_length=None
    ):
        super().__init__(
            name, r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', required=required, default=default,
            default_factory=default_factory, min_length=min_length, max_length=max_length
        )


class IntColumn(ExcelColumn):

    def __init__(
            self, name: str, /, *, required=False, default=None, default_factory=None, max_value: int | None = None,
            min_value: int | None = None, **kwargs
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, max_value=max_value,
            min_value=min_value, **kwargs
        )

    def validate(self, value):
        if value is None:
            value = self.default_factory()
        if self.required and value is None:
            raise ValueError(f'字段 {self.name} 不能为空')
        if value is None:
            return value
        try:
            value = int(value)
        except ValueError:
            raise ValueError(f'{self.name}必须为数字（实际为{value}）')
        else:
            if self.kwargs.get('max_value') is not None and value > self.kwargs.get('max_value', 0):
                raise ValueError(f'{self.name} 不能大于 {self.kwargs.get("max_value")}（实际为{value}）')
            if self.kwargs.get('min_value') is not None and value < self.kwargs.get('min_value', 0):
                raise ValueError(f'{self.name} 不能小于 {self.kwargs.get("min_value")}（实际为{value}）')
            return value

    def get_data_validation(self, columnIndex):
        operator: Literal[
            "between", "notBetween", "equal", "notEqual",
            "lessThan", "lessThanOrEqual", "greaterThan",  "greaterThanOrEqual"
        ]
        if self.kwargs.get('max_value') is not None and self.kwargs.get('min_value') is not None:
            operator = 'between'
            formula = dict(formula1=self.kwargs.get('min_value'), formula2=self.kwargs.get('max_value'),)
            errorMessage = f'{self.name}只能输入{self.kwargs.get("min_value")}~{self.kwargs.get("max_value")}之间的整数'
        elif self.kwargs.get('min_value') is not None:
            operator = 'greaterThanOrEqual'
            formula = {"formula1": self.kwargs.get('min_value')}
            errorMessage = f"{self.name}不能小于{self.kwargs.get('min_value')}"
        elif self.kwargs.get('max_value') is not None:
            operator = 'lessThanOrEqual'
            formula = {"formula1": self.kwargs.get('max_value')}
            errorMessage = f"{self.name}不能大于{self.kwargs.get('max_value')}"
        else:
            return None
        dv = DataValidation(
            type='whole',
            operator=operator,
            **formula,
            showErrorMessage=True,
            showInputMessage=False
        )
        dv.error = errorMessage
        dv.errorTitle = "无效的输入"
        columnName = get_column_letter(columnIndex)
        dv.add(f'{columnName}{2}:{columnName}{10000}')  # 数据验证区域,2-11行
        return dv


class FloatColumn(ExcelColumn):

    def __init__(
            self, name: str, /, *, required=False, default=None, default_factory=None, max_value: float = None,
            min_value: float | None = None, **kwargs
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, max_value=max_value,
            min_value=min_value, **kwargs
        )

    def validate(self, value):
        if value is None:
            value = self.default_factory()
        if self.required and value is None:
            raise ValueError(f'字段 {self.name} 不能为空')
        if value is None:
            return value
        try:
            value = float(value)
        except ValueError:
            raise ValueError(f'{self.name}必须为数字（实际为{value}）')
        else:
            if self.kwargs.get('max_value') is not None and value > self.kwargs.get('max_value'):
                raise ValueError(f'{self.name} 不能大于 {self.kwargs.get("max_value")}（实际为{value}）')
            if self.kwargs.get('min_value') is not None and value < self.kwargs.get('min_value'):
                raise ValueError(f'{self.name} 不能小于 {self.kwargs.get("min_value")}（实际为{value}）')
            return value

    def get_data_validation(self, columnIndex):
        operator: Literal[
            "between", "notBetween", "equal", "notEqual",
            "lessThan", "lessThanOrEqual", "greaterThan",  "greaterThanOrEqual"
        ]
        if self.kwargs.get('max_value') is not None and self.kwargs.get('min_value') is not None:
            operator = 'between'
            formula = dict(formula1=self.kwargs.get('min_value'), formula2=self.kwargs.get('max_value'),)
            errorMessage = f'{self.name}只能输入{self.kwargs.get("min_value")}~{self.kwargs.get("max_value")}之间的数字'
        elif self.kwargs.get('min_value') is not None:
            operator = 'greaterThanOrEqual'
            formula = {"formula1": self.kwargs.get('min_value')}
            errorMessage = f"{self.name}不能小于{self.kwargs.get('min_value')}"
        elif self.kwargs.get('max_value') is not None:
            operator = 'lessThanOrEqual'
            formula = {"formula1": self.kwargs.get('max_value')}
            errorMessage = f"{self.name}不能大于{self.kwargs.get('max_value')}"
        else:
            return None
        dv = DataValidation(
            type='decimal',
            operator=operator,
            **formula,
            showErrorMessage=True,
            showInputMessage=False
        )
        dv.error = errorMessage
        dv.errorTitle = "无效的输入"
        columnName = get_column_letter(columnIndex)
        dv.add(f'{columnName}{2}:{columnName}{10000}')  # 数据验证区域,2-11行
        return dv

class DateColumn(ExcelColumn):
    def __init__(
            self, name: str, /, *, required=False, default=None, default_factory=None, max_date: datetime.date = None,
            min_date: datetime.date = None, **kwargs
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, max_date=max_date,
            min_date=min_date, **kwargs
        )

    def validate(self, value):
        if value is None:
            value = self.default_factory()
        if self.required and value is None:
            raise ValueError(f'字段 {self.name} 不能为空')
        if value is None:
            return value
        try:
            if isinstance(value, datetime.datetime):
                value = value.date()
            elif isinstance(value, datetime.date):
                value = value
            elif isinstance(value, str):
                value = parser.parse(value).date()
            else:
                raise ValueError(f'{self.name}必须为日期格式（实际为{value}）')
        except parser.ParserError:
            raise ValueError(f'{self.name}必须为日期（实际为{value}）')
        else:
            if self.kwargs.get('max_date') is not None and value > self.kwargs.get('max_date'):
                raise ValueError(f'{self.name} 不能大于 {self.kwargs.get("max_date")}（实际为{value}）')
            if self.kwargs.get('min_date') is not None and value < self.kwargs.get('min_date'):
                raise ValueError(f'{self.name} 不能小于 {self.kwargs.get("min_date")}（实际为{value}）')
            return value


class DateTimeColumn(ExcelColumn):
    def __init__(
            self, name: str, /, *, required=False, default=None, default_factory=None, max_date: datetime.date = None,
            min_date: datetime.date = None, **kwargs
    ):
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, max_date=max_date,
            min_date=min_date, **kwargs
        )

    def validate(self, value):
        if value is None:
            value = self.default_factory()
        if self.required and value is None:
            raise ValueError(f'字段 {self.name} 不能为空')
        if value is None:
            return value
        try:
            if isinstance(value, datetime.datetime):
                value = value
            elif isinstance(value, datetime.date):
                value = datetime.datetime.combine(value, datetime.datetime.min.time())
            elif isinstance(value, str):
                value = parser.parse(value)
            else:
                raise ValueError(f'{self.name}必须为日期时间格式（实际为{value}）')
        except parser.ParserError:
            raise ValueError(f'{self.name}必须为日期（实际为{value}）')
        else:
            if self.kwargs.get('max_date') is not None and value > self.kwargs.get('max_date'):
                raise ValueError(f'{self.name} 不能大于 {self.kwargs.get("max_date")}（实际为{value}）')
            if self.kwargs.get('min_date') is not None and value < self.kwargs.get('min_date'):
                raise ValueError(f'{self.name} 不能小于 {self.kwargs.get("min_date")}（实际为{value}）')
            return value


class StringSequenceColumn(StringColumn):
    def __init__(self, name: str, sequence: dict[Any, str] | list[str], /, *, required=False, default=None,
                 default_factory=None,
                 **kwargs):
        if isinstance(sequence, list):
            sequence = {v: v for v in sequence}
        sequence_reverse = {v: k for k, v in sequence.items()}
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, sequence=sequence,
            sequence_reverse=sequence_reverse, **kwargs)

    def validate(self, value):
        value = super().validate(value)
        if value is None:
            return value
        if value not in self.kwargs['sequence_reverse']:
            raise ValueError(f'{self.name}必须在{'/'.join(map(str, self.kwargs["sequence_reverse"].keys()))}中')
        return self.kwargs['sequence_reverse'][value]

    def get_data_validation(self, columnIndex):
        dv = DataValidation(
            type='list',
            formula1=f'"{",".join(self.kwargs['sequence'].values())}"',
            allow_blank=not self.required,
            showErrorMessage=True,
            showInputMessage=False
        )
        dv.error = f"{self.name}只能从【{'、'.join(self.kwargs['sequence'].values())}】中选择"
        dv.errorTitle = "无效的输入"
        dv.prompt = f'请从【{'、'.join(self.kwargs['sequence'].values())}】中选择数据'
        dv.promptTitle = f"选择{self.name}"
        columnName = get_column_letter(columnIndex)
        dv.add(f'{columnName}{2}:{columnName}{10000}')  # 数据验证区域,2-11行
        return dv


class IntSequenceColumn(IntColumn):
    def __init__(self, name: str, sequence: dict[Any, int] | list[int], /, *, required=False, default=None,
                 default_factory=None,
                 **kwargs):
        if isinstance(sequence, list):
            sequence = {v: v for v in sequence}
        sequence_reverse = {v: k for k, v in sequence.items()}
        super().__init__(
            name, required=required, default=default, default_factory=default_factory, sequence=sequence,
            sequence_reverse=sequence_reverse, **kwargs)

    def validate(self, value):
        value = super().validate(value)
        if value is None:
            return value
        if value not in self.kwargs['sequence_reverse']:
            raise ValueError(f'{self.name}必须在{'/'.join(map(str, self.kwargs["sequence_reverse"].keys()))}中')
        return self.kwargs['sequence_reverse'][value]

    def get_data_validation(self, columnIndex):
        dv = DataValidation(
            type='list',
            formula1=f'"{",".join(map(str, self.kwargs['sequence'].values()))}"',
            allow_blank=not self.required,
            showErrorMessage=True,
            showInputMessage=False
        )
        dv.error = f"{self.name}只能从【{'、'.join(self.kwargs['sequence'].values())}】中选择"
        dv.errorTitle = "无效的输入"
        dv.prompt = f'请从【{'、'.join(self.kwargs['sequence'].values())}】中选择数据'
        dv.promptTitle = f"选择{self.name}"
        columnName = get_column_letter(columnIndex)
        dv.add(f'{columnName}{2}:{columnName}{10000}')  # 数据验证区域,2-11行
        return dv


SequenceColumn = StringSequenceColumn

if __name__ == '__main__':
    pass
