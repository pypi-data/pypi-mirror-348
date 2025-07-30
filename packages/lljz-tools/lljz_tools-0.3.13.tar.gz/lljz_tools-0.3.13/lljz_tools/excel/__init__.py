from os import PathLike
from io import BytesIO
import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..excel.reader import ExcelReader
from ..excel.writer import ExcelWriter


class Excel(ExcelReader, ExcelWriter):

    def __init__(self, filename: str | PathLike | BytesIO):
        super().__init__(filename=filename)
        self._sheets = {name: self._excel[name] for name in self._excel.sheetnames}  

    def _get_excel(self):  # noqa
        if isinstance(self._file, str) or isinstance(self._file, PathLike):
            if not os.path.exists(self._file):
                return Workbook()
            else:
                return load_workbook(self._file, data_only=True)
        else:
            return load_workbook(self._file, data_only=True)

    def write(self, data, /, *, sheet_name=None):
        if not sheet_name:
            sheet_name = self._excel.active.title  # type: ignore
        return super().write(data, sheet_name=sheet_name)

    def __getitem__(self, item) -> Worksheet:
        if not self._excel:
            raise ValueError('excel not open')
        return self._excel[item]  # type: ignore
