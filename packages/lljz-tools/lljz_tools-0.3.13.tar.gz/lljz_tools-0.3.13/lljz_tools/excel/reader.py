import csv
from datetime import datetime
from io import BytesIO
from os import PathLike
from typing import Generator, overload, Literal, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _read_by_range(data: Iterable, min_range: int, max_range: int):
    """
    只读取范围内的数据
    :param data:
    :param min_range: 从1开始，包含
    :param max_range: 包含
    :return:
    """
    for i, val in enumerate(data, start=1):
        if i < min_range:
            continue
        if i > max_range:
            break
        yield val

class CSVReader:
    def __init__(self, filename: str | PathLike):
        self._file = filename
    
    @overload
    def read(self, has_title: Literal[True] = True, min_row: int | None =None, max_row: int | None =None, min_col: int | None =None, max_col: int | None =None) -> Generator[dict[str, str], None, None]:
        ...
    
    @overload
    def read(self, has_title: Literal[False] = False, min_row: int | None =None, max_row: int | None =None, min_col: int | None =None, max_col: int | None =None) -> Generator[list[str], None, None]:
        ...
    
    def read(self, has_title: bool = True, min_row: int | None =None, max_row: int | None =None, min_col: int | None =None, max_col: int | None =None):
        min_row = min_row or 1
        max_row = max_row or 999999999
        min_col = min_col or 1
        max_col = max_col or 999999999
        assert 1 <= min_row <= max_row <= 999999999 and 1 <= min_col <= max_col <= 999999999, \
            "参数有误，行数范围限制为[1, 999999999]"

        with open(self._file) as f:
            reader = (_read_by_range(row, min_col, max_col) for row in _read_by_range(csv.reader(f), min_row, max_row))

            if has_title:
                title = next(reader)

                yield from (dict(zip(title, row)) for row in reader)
            else:
                yield from reader


ExcelData = str | float | datetime | None


class ExcelReader:

    def __init__(self, filename: str | PathLike | BytesIO):
        self._file = filename
        if isinstance(self._file, str) and self._file.endswith('.csv'):
            raise ValueError('ExcelReader不支持csv文件，请使用CSVReader')
        self._excel = self._get_excel()

    def _get_excel(self):
        return load_workbook(self._file, read_only=True, data_only=True)

    def _get_sheet(self, sheet: int | None | str) -> Worksheet:
        if isinstance(sheet, int):
            return self._excel.worksheets[sheet]
        elif isinstance(sheet, str):
            return self._excel[sheet]
        s = self._excel.active  
        if not s:
            raise ValueError('Excel不包含默认激活的Sheet，请指定一个Sheet名称或者索引')
        return s
    
    @overload
    def read(self, __sheet: int | None | str = None, /, *, has_title: Literal[True] = True, min_row=None, max_row=None,
             min_col=None, max_col=None) -> Generator[dict, None, None]:
        ...
    
    @overload
    def read(self, __sheet: int | None | str = None, /, *, has_title: Literal[False] = False, min_row=None, max_row=None,
             min_col=None, max_col=None) -> Generator[tuple, None, None]:
        ...

    def read(self, __sheet: int | None | str = None, /, *, has_title: bool = True, min_row=None, max_row=None,
             min_col=None, max_col=None):
        """
        读取Excel的Sheet页中的数据

        如果没有指定索引，则范围从A1开始。如果工作表中没有单元格，则返回一个空元组。

        :param __sheet: sheet页的顺序或名字或默认
        :type __sheet: int | None | str

        :param has_title: 是否包含标题，如果包含标题则每行数据都会以字典的形式返回，否则以元组的形式返回
        :type has_title: bool

        :param min_col: 最小列索引(从1开始)
        :type min_col: int

        :param min_row: 最小行索引(从1开始)
        :type min_row: int

        :param max_col: 最大列索引(从1开始)
        :type max_col: int

        :param max_row: 最大行索引(从1开始)
        :type max_row: int

        :rtype: generator
        """
        sheet = self._get_sheet(__sheet)
        iter_rows = sheet.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
        if has_title:
            title = next(iter_rows)
            yield from (dict(zip(title, r)) for r in iter_rows)
        else:
            yield from iter_rows

    @property
    def sheet_names(self):
        return self._excel.sheetnames

    def close(self):
        self._excel.close()
