# coding=utf-8

"""
@fileName       :   template.py
@data           :   2024/11/15
@author         :   jiangmenggui@hosonsoft.com
"""
from typing import NamedTuple

from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import PatternFill, Color
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lljz_tools.excel import ExcelReader
from .columns import *


class _ExcelTemplateMeta(type):

    def __new__(cls, name, bases, attrs, total=False, variable=True, check_type=False):
        if name in ['ExcelTemplate']:
            return type.__new__(cls, name, bases, attrs)
        __columns__ = {}
        __columns_mapper__ = {}
        __columns_mapper_reverse__ = {}

        for k, v in attrs.items():
            if not isinstance(v, ExcelColumn):
                continue
            __columns__[k] = v
            name = f'*{v.name}' if v.required else v.name
            __columns_mapper__[k] = name
            __columns_mapper_reverse__[name] = k
        attrs['__columns__'] = __columns__
        attrs['__columns_mapper__'] = __columns_mapper__
        attrs['__columns_mapper_reverse__'] = __columns_mapper_reverse__
        return type.__new__(cls, name, bases, attrs)


class ExcelConfig(NamedTuple):
    # sheet页名称，为None则读取默认页
    sheet_name: str = None
    # 唯一的数据列
    unique_columns: list = []
    # 允许空模板
    allow_empty_template: bool = True
    # 忽略空行
    ignore_empty_row: bool = True
    # Excel起始行
    start_row: int = 1
    # Excel最多读取的行
    end_row: int = None


class ExcelTemplate(metaclass=_ExcelTemplateMeta):
    config = ExcelConfig()

    @classmethod
    def _validate_data(cls, data: list[dict]):
        newData = []
        for row in data:
            row = {k: row.get(v) for k, v in cls.__columns_mapper__.items()}
            newRow = {}
            if not cls.config.ignore_empty_row and not any(bool(val) for val in row.values()):
                raise ValueError('数据不能为空')
            for key, value in row.items():
                if key not in cls.__columns__:
                    continue
                column = cls.__columns__[key]
                newRow[key] = column.validate(value)
            newData.append(newRow)
        if cls.config.unique_columns:
            unique_columns = [tuple(row.get(column) for column in cls.config.unique_columns) for row in newData]
            _data = {}
            for rowNo, row in enumerate(unique_columns, start=2):
                if row in _data:
                    raise ValueError(f'与第{_data[row]}行数据重复')
                _data[row] = rowNo
        if not cls.config.allow_empty_template and not newData:
            raise ValueError('模板不能为空')
        return newData

    @classmethod
    def read_excel(cls, file):
        excel = ExcelReader(file)
        data = excel.read(cls.config.sheet_name, min_row=cls.config.start_row, max_row=cls.config.end_row)
        return cls._validate_data(data)

    @staticmethod
    def _write_title(sheet: Worksheet, columns: dict[str, ExcelColumn]):

        title_fill = PatternFill('solid', fgColor='00B0F0')

        required = TextBlock(InlineFont(rFont='等线', sz=12, b=True, color=Color('fa201c')), '*')
        required.name = '*'
        title_font = InlineFont(rFont='等线', sz=12, b=True)

        def get_title_cell(column: ExcelColumn):
            if column.required:
                return CellRichText(
                    required,
                    TextBlock(title_font, column.name),
                )
            value = CellRichText(TextBlock(title_font, column.name))
            cell = WriteOnlyCell(sheet, value=value)  # noqa
            cell.fill = title_fill
            return cell

        row = [get_title_cell(column) for column in columns.values()]
        sheet.append(row)

    @classmethod
    def create_excel_template(cls, file):
        wb = Workbook(write_only=True)
        sheet: Worksheet = wb.create_sheet(title=cls.config.sheet_name or 'sheet1')
        cls._write_title(sheet, cls.__columns__)
        # 设置数据验证
        for index, (k, v) in enumerate(cls.__columns__.items(), start=1):
            dv = v.get_data_validation(columnIndex=index)
            if dv:
                sheet.data_validations.append(dv)
        wb.save(file)


if __name__ == '__main__':
    pass
