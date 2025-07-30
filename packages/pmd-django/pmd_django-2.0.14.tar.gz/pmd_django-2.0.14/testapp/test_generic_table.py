import io
import json
from datetime import datetime, timezone

from django.db.models import Count
from django.http import FileResponse, JsonResponse
from django.test import RequestFactory, TestCase
from django.utils.timezone import now
from openpyxl.reader.excel import load_workbook

from pmd_django import GenericTableView
from pmd_django.generic_table.generic_table import (
    _apply_filters,
    _apply_ordering,
    _apply_user_filters,
    _build_counts,
    _generate_download_response,
    _get_allowed_fields,
    _get_page_size,
    _paginate_queryset,
    view,
)
from testapp.models import TestRelatedModel, TestModel

class TestGenericTable(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.factory = RequestFactory()

        tm = TestModel.objects.create(status="active")

        TestModel.objects.bulk_create(
            [
                TestModel(status="inactive"),
                TestModel(status="active"),
                TestModel(status="inactive"),
                TestModel(status="active"),
            ]
        )

        TestRelatedModel.objects.create(name="Related Item", related=tm)

    def test_build_counts(self):
        counted_values = ["active", "inactive", "pending"]
        queryset = TestModel.objects.all()
        result = _build_counts(queryset, "status", counted_values)

        self.assertEqual(result, {"active": 3, "inactive": 2, "pending": 0, "ALL": 5})

    def test_get_page_size_valid(self):
        request = self.factory.get("/?page_size=20")
        self.assertEqual(_get_page_size(request, None), 20)

    def test_get_page_size_invalid(self):
        request = self.factory.get("/?page_size=invalid")
        self.assertEqual(_get_page_size(request, None), 30)

    def test_apply_filters(self):
        request = self.factory.get("/?status=active")
        extra_filters = {}
        extra_filter_from_request = {"status": "status__iexact"}

        filtered_qs = _apply_filters(
            TestModel.objects.all(), request, extra_filters, extra_filter_from_request
        )
        self.assertTrue(filtered_qs.exists())

    def test_paginate_queryset(self):
        request = self.factory.get("/?page=2&page_size=2")
        queryset = TestModel.objects.all().order_by("id")

        page_obj, paginator = _paginate_queryset(queryset, request, 2)

        self.assertEqual(page_obj.number, 2)
        self.assertEqual(paginator.num_pages, 3)

        all_ids = list(TestModel.objects.all().order_by("id").values_list("id", flat=True))
        expected_ids = all_ids[2:4]  # page 2, page_size 2
        actual_ids = list(page_obj.object_list.values_list("id", flat=True))

        self.assertEqual(actual_ids, expected_ids)

    def test_view(self):
        request = self.factory.get("/?page=1&sort_by=id")
        queryset = TestModel.objects.all()

        response = view(queryset, request, "status", ["active", "inactive"])

        self.assertIsInstance(response, JsonResponse)
        response_data = json.loads(response.content)

        self.assertIn("pagination", response_data)
        self.assertIn("stageCounts", response_data)

    def test_valid_user_filters(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [
                        {"field": "status", "condition": "exact", "value": "active"},
                        {
                            "field": "id",
                            "condition": "gte",
                            "value": 3,
                            "logicalOperator": "AND",
                        },
                    ]
                )
            },
        )
        queryset = _apply_user_filters(TestModel.objects.all(), request)
        self.assertTrue(queryset.exists())

    def test_invalid_field_filtering(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [{"field": "password", "condition": "exact", "value": "test"}]
                )
            },
        )
        with self.assertRaises(ValueError) as e:
            _apply_user_filters(TestModel.objects.all(), request)
        self.assertEqual(str(e.exception), "Filtering on 'password' is not allowed")

    def test_invalid_condition_filtering(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [{"field": "status", "condition": "hacked", "value": "active"}]
                )
            },
        )
        with self.assertRaises(ValueError) as e:
            _apply_user_filters(TestModel.objects.all(), request)
        self.assertEqual(str(e.exception), "Condition 'hacked' is not supported")

    def test_json_parsing_error(self):
        request = self.factory.get("/", data={"user_filters": "[INVALID JSON"})
        with self.assertRaises(ValueError) as e:
            _apply_user_filters(TestModel.objects.all(), request)
        self.assertEqual(str(e.exception), "Invalid JSON in user_filters")

    def test_missing_value_filter(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [{"field": "status", "condition": "exact", "value": None}]
                )
            },
        )
        queryset = _apply_user_filters(TestModel.objects.all(), request)
        self.assertEqual(queryset.count(), 5)  # Should return full dataset

    def test_apply_user_filters_where_we_have_dangling_operator(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [
                        {
                            "field": "id",
                            "condition": "gte",
                            "value": 3,
                            "logicalOperator": "AND",
                        }
                    ]
                )
            },
        )

        filtered_qs = _apply_user_filters(TestModel.objects.all(), request)
        self.assertTrue(filtered_qs.exists())

    def test_sql_injection_attempt(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [
                        {
                            "field": "status",
                            "condition": "exact",
                            "value": "'; DROP TABLE users; --",
                        }
                    ]
                )
            },
        )
        queryset = _apply_user_filters(TestModel.objects.all(), request)
        self.assertEqual(queryset.count(), 0)  # Should not execute harmful queries

    def test_unmatched_operators_and_filters(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [
                        {"field": "status", "condition": "exact", "value": "active"},
                        {
                            "field": "id",
                            "condition": "gte",
                            "value": 3,
                            "logicalOperator": "AND",
                        },
                        {
                            "field": "status",
                            "condition": "icontains",
                            "value": "inactive",
                        },
                        # No logicalOperator for this one
                    ]
                )
            },
        )
        with self.assertRaises(ValueError) as e:
            _apply_user_filters(TestModel.objects.all(), request)
        self.assertIn(
            "Filter-operator mismatch: expected 2 operators, got 1", str(e.exception)
        )

    def test_standard_model_fields(self):
        qs = TestModel.objects.all()
        allowed_fields = _get_allowed_fields(qs)

        self.assertIn("id", allowed_fields)
        self.assertIn("status", allowed_fields)

    def test_related_model_fields(self):
        qs = TestModel.objects.all()
        allowed_fields = _get_allowed_fields(qs)

        self.assertIn("testrelatedmodel__name", allowed_fields)

    def test_annotated_fields(self):
        qs = TestModel.objects.annotate(status_count=Count("status"))
        allowed_fields = _get_allowed_fields(qs)

        self.assertIn("status_count", allowed_fields)

    def test_values_fields(self):
        qs = TestModel.objects.values("id", "status")
        qs.extra_allowed_fields = {"id", "status"}

        allowed_fields = _get_allowed_fields(qs)

        self.assertIn("id", allowed_fields)
        self.assertIn("status", allowed_fields)
        self.assertNotIn("nope", allowed_fields)

    def test_rejects_unauthorized_fields(self):
        qs = TestModel.objects.all()
        allowed_fields = _get_allowed_fields(qs)

        self.assertNotIn("password_hash", allowed_fields)  # Doesn't exist in the model
        self.assertNotIn("random_field", allowed_fields)  # Should not be allowed

    def test_apply_user_filters_with_annotated_field(self):
        qs = TestModel.objects.annotate(status_count=Count("status"))

        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [{"field": "status_count", "condition": "gte", "value": 1}]
                )
            },
        )
        filtered_qs = _apply_user_filters(qs, request)
        self.assertTrue(filtered_qs.exists())

    def test_apply_user_filters_with_related_field(self):
        request = self.factory.get(
            "/",
            data={
                "user_filters": json.dumps(
                    [
                        {
                            "field": "testrelatedmodel__name",
                            "condition": "exact",
                            "value": "Related Item",
                        }
                    ]
                )
            },
        )
        qs = _apply_user_filters(TestModel.objects.all(), request)
        self.assertTrue(qs.exists())

    def test_generate_excel_response(self):
        queryset = TestModel.objects.all()
        values_list = ["status"]

        response = _generate_download_response(queryset, values_list)
        worksheet = self._load_excel_from_response(response)

        expected_headers = ["status"]
        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, expected_headers)

        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        expected_data = [
            ("active",),
            ("inactive",),
            ("active",),
            ("inactive",),
            ("active",),
        ]
        self.assertEqual(data_rows, expected_data)

    def test_generate_excel_response_with_list_field(self):
        TestModel.objects.create(status="tagged", tags=["TEST", "ALPHA"])

        queryset = TestModel.objects.filter(status="tagged")
        values_list = ["status", "tags"]

        response = _generate_download_response(queryset, values_list)
        worksheet = self._load_excel_from_response(response)

        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, ["status", "tags"])

        data_row = list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(data_row[0], "tagged")
        self.assertIsInstance(data_row[1], str)
        self.assertIn("TEST", data_row[1])
        self.assertIn("ALPHA", data_row[1])

    def test_generate_excel_response_with_timezone_datetime(self):
        aware_datetime = now().astimezone(timezone.utc)
        TestModel.objects.create(status="datetime", created_at=aware_datetime)

        queryset = TestModel.objects.filter(status="datetime")
        values_list = ["status", "created_at"]

        response = _generate_download_response(queryset, values_list)
        worksheet = self._load_excel_from_response(response)

        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, ["status", "created_at"])

        data_row = list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(data_row[0], "datetime")
        self.assertIsInstance(data_row[1], datetime)
        self.assertIsNone(data_row[1].tzinfo, "Datetime should be naive (tzinfo=None)")

    def test_json_and_excel_field_handling(self):
        class DummyView(GenericTableView):
            model = TestModel
            stage_field = "status"
            table_headers = {
                "id": None,
                "status": ("Status", ("default", {})),
                "created_at": None,
            }

            @property
            def generic_table_view_kwargs(self):
                return {
                    "field": self.stage_field,
                    "values_list": list(self.table_headers.keys()),
                    "download_fields": [k for k, v in self.table_headers.items() if v is not None or k == "id"],
                    "counted_values": [],
                    "data_key": "data",
                    "final_json_hook": self.final_json_hook,
                }

            def final_json_hook(self, data):
                data["displayHeaders"] = {k: v[0] for k, v in self.table_headers.items() if v}
                data["filterFieldTypes"] = {
                    k: v[1][0] for k, v in self.table_headers.items() if v
                }
                return data

        # Excel export should only include 'id' and 'status'
        excel_request = self.factory.get("/", HTTP_ACCEPT="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        view_instance = DummyView()
        excel_response = view_instance.get(excel_request)
        worksheet = self._load_excel_from_response(excel_response)
        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, ["id", "status"])

        # JSON export should include all table_headers keys (even those with None)
        json_request = self.factory.get("/")
        json_response = view_instance.get(json_request)
        data = json.loads(json_response.content)["data"][0]
        self.assertIn("id", data)
        self.assertIn("status", data)
        self.assertIn("created_at", data)

    def _load_excel_from_response(self, response: FileResponse):
        self.assertIsInstance(response, FileResponse)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        output = io.BytesIO()
        for chunk in response.streaming_content:
            output.write(chunk)
        output.seek(0)
        workbook = load_workbook(output)
        return workbook.active

    def test_apply_ordering_id(self):
        request = self.factory.get("/?sort_by=id&sort_order=desc")
        ordered_qs = _apply_ordering(TestModel.objects.all(), request)
        ids = list(ordered_qs.values_list("id", flat=True))
        self.assertEqual(ids, sorted(ids, reverse=True))

    def test_apply_ordering_case_insensitive_for_strings(self):
        TestModel.objects.create(status="Beta")
        TestModel.objects.create(status="alpha")
        TestModel.objects.create(status="Gamma")

        request = self.factory.get("/?sort_by=status&sort_order=asc")
        ordered_qs = _apply_ordering(TestModel.objects.all(), request)
        ordered_statuses = list(ordered_qs.values_list("status", flat=True))

        self.assertEqual(ordered_statuses, sorted(ordered_statuses, key=lambda s: s.lower()))