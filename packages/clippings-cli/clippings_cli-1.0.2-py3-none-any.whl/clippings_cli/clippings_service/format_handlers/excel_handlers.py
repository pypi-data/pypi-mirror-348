"""
File containing functions for handling Excel Clippings file.
"""

import os
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

FIELDS: OrderedDict[str, dict] = OrderedDict(
    [
        ("Book title", {"fetch_method": lambda clipping: clipping["book"]["title"], "width": 20}),
        ("Book author", {"fetch_method": lambda clipping: clipping["book"]["author"], "width": 20}),
        ("Content", {"fetch_method": lambda clipping: clipping["content"], "width": 100}),
        ("Page number", {"fetch_method": lambda clipping: clipping["page_number"], "width": 10, "is_number": True}),
        ("Location", {"fetch_method": lambda clipping: clipping["location"], "width": 10}),
        ("Created at", {"fetch_method": lambda clipping: clipping["created_at"], "width": 10}),
        ("Clipping type", {"fetch_method": lambda clipping: clipping["clipping_type"], "width": 10}),
        ("Errors", {"fetch_method": lambda clipping: str(clipping["errors"]), "width": 20}),
    ]
)
HEADERS_STYLING = {
    "font": Font(bold=True, color="FFFFFF"),
    "fill": PatternFill(start_color="595959", end_color="595959", fill_type="solid"),
    "alignment": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    ),
}

DATA_STYLING: dict = {
    "font": Font(color="000000"),
    "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center"),
    "border": Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    ),
}


def apply_headers_styling(ws: Worksheet) -> None:
    """
    Function to apply styling to Worksheet headers:
    * Background color
    * Font color
    * Borders
    * Enable filtering
    * Column width

    Args:
        ws (Worksheet): Excel worksheet object.
    """
    for cell in ws[1]:
        cell.font = HEADERS_STYLING["font"]
        cell.fill = HEADERS_STYLING["fill"]
        cell.alignment = HEADERS_STYLING["alignment"]
        cell.border = HEADERS_STYLING["border"]

    # Add filters to the headers
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = FIELDS[col[0].value].get("width", 2)


def apply_data_cells_styling(ws: Worksheet):
    """
    Function to apply styling to Worksheet data cells:
    * Background color
    * Font color
    * Borders

    Args:
        ws (Worksheet): Excel worksheet object.
    """
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = DATA_STYLING["font"]
            cell.fill = DATA_STYLING["fill"]
            cell.alignment = DATA_STYLING["alignment"]
            cell.border = DATA_STYLING["border"]


def generate_excel(clippings: list[dict[str, Any]], output_path: Path | str) -> dict:
    """
    In provided output_path creates Excel file containing data collected from Clippings input file.

    Args:
        clippings (list[dict]): List of collected Clippings.
        output_path (Path | str): Path to output file.

    Returns:
        dict: Dictionary containing data about potential errors.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(list(FIELDS.keys()))

    for clipping in clippings:
        ws.append([FIELDS[key]["fetch_method"](clipping) for key in FIELDS])

    apply_headers_styling(ws)
    apply_data_cells_styling(ws)

    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
    except PermissionError as e:
        return {"error": e}
    return {}
