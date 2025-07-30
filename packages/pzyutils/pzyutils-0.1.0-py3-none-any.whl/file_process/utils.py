import os
import json
import shutil
import openpyxl
import traceback
import platform
import sys


def read_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as fr:
        return json.load(fr)

def get_file_path_split():
    system_name = platform.system()
    return "\\" if system_name == "Windows" else "/"


def write_json(file_path, data_dict, indent=None):
    with open(file_path, 'w', encoding='utf-8') as fw:
        fw.write(json.dumps(data_dict, ensure_ascii=False, indent=indent))
    print('write json ok. {}'.format(file_path))


def read_json_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as fr:
        return [json.loads(x.strip()) for x in fr]


def write_json_list(file_path, data_list):
    with open(file_path, 'w', encoding='utf-8') as fw:
        [fw.write(json.dumps(x, ensure_ascii=False) + '\n') for x in data_list]
    print('write json list ok. {}'.format(file_path))


def read_text_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as fr:
        return [x.strip() for x in fr]


def write_text_list(file_path, data_list):
    with open(file_path, 'w', encoding='utf-8') as fw:
        [fw.write('{}\n'.format(x)) for x in data_list]
    print('write txt ok. {}'.format(file_path))


def read_xlsx(file_path):
    # print('start read xlsx. {}'.format(file_path))
    wb = openpyxl.load_workbook(file_path)
    sheet_name_list = wb.sheetnames
    sheet_list = []
    for sheet_name in sheet_name_list:
        data_list = list(wb[sheet_name].values)
        sheet_list.append({'sheet_name': sheet_name, 'header': data_list[0], 'data': data_list[1:]})
    # print('read xlsx ok. {}'.format(file_path))
    return sheet_list


def write_xlsx(file_path, all_sheet_list):
    print('start write xlsx. {}'.format(file_path))
    wb = openpyxl.Workbook(write_only=True)
    for sheet_dict in all_sheet_list:
        ws = wb.create_sheet(title=sheet_dict['sheet_name'])
        ws.append(sheet_dict['header'])
        [ws.append(sub_list) for sub_list in sheet_dict['data']]
    wb.save(file_path)
    print('write xlsx ok. {}'.format(file_path))


def delete_file(file_path):
    try:
        if not os.path.exists(file_path):
            return False
        if os.path.isfile(file_path):  # file
            os.remove(file_path)
        else:  # dir
            shutil.rmtree(file_path)
        return True
    except:
        print('file: {}, msg: {}'.format(file_path, traceback.format_exc()))
        status = os.system('rm -rf {}'.format(file_path))
        return status == 0