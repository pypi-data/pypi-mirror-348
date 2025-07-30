from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from typing import List
from io import BytesIO


def iniciar_planilha(conteudo):
    """
    Inicia o arquivo Excel contendo os dados.
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :return: Retorna a planilha.
    """
    try:
        planilha = load_workbook(filename=conteudo, data_only=True)
    except PermissionError:
        print('O arquivo já está aberto, feche o mesmo antes de prosseguir.')
        exit()
    else:
        return planilha


def descobrir_linha_vazia_planilha(conteudo, coluna: str) -> str:
    """
    Descobre o número da ultima linha preenchida na planilha da coluna informada.
    Mesmo se houver linhas vazias entres linhas preenchidas, a última será pega.
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :param coluna: Coluna da planilha onde será procurado o último elemento.
    :return: Retorna o número da ultima linha como uma string.
    """
    planilha = iniciar_planilha(conteudo)
    aba_ativa = planilha.active

    ultima_posicao: int = aba_ativa[coluna][-1].row

    ultima_linha: int = 1
    cont: int = 0

    for i, celula in enumerate(aba_ativa[f'{coluna}1:{coluna}{ultima_posicao}']):
        if celula[0].value is not None:
            ultima_linha: int = celula[0].row
        else:
            # Ignoro a quantidade de linhas que eu percebi que eu preciso passar para chegar na proxima linha NÃO vazia:
            if cont > 0:
                cont -= 1  # Diminuo o cont pois acabei de passar por uma linha que estou ignorando.
            else:
                # aba_ativa[coluna][i].row == celula[0].row  | Ambos são a mesma coisa.
                if aba_ativa[coluna][i].row < ultima_posicao:
                    # VERIFICANDO SE EXISTE ALGUMA LINHA PREENCHIDA DA POSIÇÃO ATUAL ATE A ULTIMA LINHA DA COLUNA:
                    quantidade_linha_faltante: int = ultima_posicao - celula[0].row
                    elemento: int = 0  # -> Variável utilizada para verificar se foi encontrada uma linha NÃO vazia.
                    # Loop baseado na quantidade de linhas faltantes, se faltam 7 linhas para acabar,
                    # o loop vai rodar 7 vezes.
                    for c in range(1, quantidade_linha_faltante + 1):
                        valor = aba_ativa[coluna][i + c].value  # PEGANDO O VALOR PRESENTE NA LINHA *SEGUINTE*.
                        # ^ Para verificar em seguida se a mesma é uma linha vazia ou não.
                        if valor is not None:
                            elemento: int = 1  # -> Indicador de que encontrei uma linha NÃO vazia.
                            cont = c - 1  # -> Quantidade de linhas que posso ignorar até a proxima linha NÃO vazia.
                            break
                            # ^ Se encontro alguma linha não vazia no meio do caminho, nem continuo olhando os próximos,
                            # pois já achei o que queria.
                    if elemento == 0:  # Se não encontrei nenhuma linha preenchida saio do loop e já sei a última linha.
                        break
                else:
                    break

    planilha.close()

    return str(ultima_linha)


def is_merged(cell, merged_ranges):
    """ Verifica se a célula faz parte de uma faixa mesclada """
    for merged_range in merged_ranges:
        # cell.coordinate: Retorna a referência da célula (por exemplo, 'A1', 'B2', etc.).
        if cell.coordinate in merged_range:  # Testa se a referência da célula atual está dentro da faixa mesclada.
            return True
    return False


def converter_objeto_para_planilha(data: List[dict]) -> bytes:
    """
    Converte um objeto para uma planilha.
    :param data:
    :return: Retorna a planilha no formato de bytes.
    """
    # Cria um novo workbook
    wb = Workbook()
    ws = wb.active

    # Adiciona o cabeçalho ao array "headers":
    headers = list(dict(data[0]).keys()) if data else []

    # Aplica o estilo negrito da planilha em uma variável:
    bold_font = Font(bold=True)

    # Alterando a posição do campo "unitPrice" para ficar após o campo "currency":
    # headers.insert(12, 'unitPrice')
    # headers = headers[0: len(headers) - 1]

    # Transformando o formato "supplierReference" em "SUPPLIER REFERENCE":
    headers = [camel_case_to_upper_case_with_spaces(element) for element in headers]

    # Adicionando o cabeçalho na planilha:
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header.upper())
        cell.font = bold_font  # Define a celula atual da planilha como negrito.

    # Adicionando os valores nas linhas:
    for row in data:
        ws.append(list(dict(row).values()))

    # Salva o workbook em um BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)

    # Move o ponteiro de volta para o início, permitindo que o conteúdo do buffer seja lido corretamente.
    # Caso eu tentasse usar o .getvalue sem fazer o .seek(0) o valor lido seria vazio, pois o ponteiro estaria
    # no final do conteudo.
    buffer.seek(0)

    # Retorna o arquivo Excel como uma resposta
    return buffer.getvalue()


def camel_case_to_upper_case_with_spaces(s: str):
    """
    Função responsável por transformar "camelCase" para "UPPER CASE WITH SPACES".
    :param s: String que será transformada.
    :return: String transformada.
    """
    import re

    # Insere um espaço antes de cada letra maiúscula
    s_with_spaces = re.sub(r'(?<!^)(?=[A-Z])', ' ', s)
    # Converte para maiúsculas
    return s_with_spaces.upper()
