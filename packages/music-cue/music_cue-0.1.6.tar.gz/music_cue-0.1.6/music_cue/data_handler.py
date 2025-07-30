import os
import re
from datetime import timedelta
from typing import DefaultDict
from itertools import zip_longest
from dataclasses import dataclass
from collections import defaultdict

import openpyxl
from music_cue.utils import xlref, custom_layout_sheet, read_excel_tab, read_original_excel_tab


class SheetExistsException(Exception):
    pass


class DataHandler:
    """
    This class provides CRUD functionality for the data entry APP to administer
    clips being used in TV episodes. The collected data is used for copyright
    purposes

    Given an original Excel file with a source sheet, a dB sheet is generated and
    per Episode a sheet is generated with information how long clips have been used
    Also an extra sheet is present which shows what clips are present per episode.

    Given the start and end dates of the clips, it is possible to split the episode
    music file into separate music files for playback purposes.
    """
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    GENERAL_ARTIST_TITLE_OVERVIEW_NAME = 'General Artist_Title overview'
    DB_SHEET_NAME = 'Database sheet'
    FILE_TYPES = ['mp3', 'm4a', 'wav', 'mov', 'mp4', 'aif', 'aiff']

    def __init__(self):

        self.wb = None
        self.excel_filename = None
        self.project_root_dir = None
        self.source_sheet_name = None

    def read_db_sheet(self) -> list[dataclass]:
        """
        Returns DB sheet as list of data class objects
        """
        fields = [
            ('Episode name', 'episode_name'), ('Event', 'event'), ('Clip name', 'clip_name'),
            ('Artist', 'artist'), ('Title', 'title'), ('Start time', 'start'), ('End time', 'end'),
            ('Duration', 'duration')
        ]
        return read_excel_tab(self.wb, self.DB_SHEET_NAME, fields)

    def read_source_sheet(self) -> list[dataclass]:
        """
        Returns original Excel data source as list of dataclass objects
        Sheet must be cleaned before running the APP. This means that
        in the column "SESSION NAME" the session names must be populated.
        """
        return read_original_excel_tab(self.wb, self.source_sheet_name)

    def get_duration_data_per_episode_per_clip(self) -> DefaultDict[str, DefaultDict[str, timedelta]]:
        """
        Returns dict with duration data per Episode per clip.
        If clip is used more than 1 time in Episode then
        duration of individual clips are added.
        """
        duration_data = defaultdict(lambda: defaultdict(timedelta))

        for row in self.read_db_sheet():

            min_start, sec_start, _ = row.start.split(':')
            min_end, sec_end, _ = row.end.split(':')

            time_end = timedelta(hours=0, minutes=int(min_end), seconds=int(sec_end))
            time_start = timedelta(hours=0, minutes=int(min_start), seconds=int(sec_start))
            time_duration = time_end - time_start

            duration_data[row.episode_name][row.clip_name] += time_duration

        return duration_data

    def update_artist_title_info(self, clip_name: str, artist: str, title: str) -> None:
        """
        For a given clip name, artist and title entries in the DB sheet are updated.
        """
        sheet = self.wb[self.DB_SHEET_NAME]

        for i, row in enumerate(self.read_db_sheet()):
            if row.clip_name != clip_name:
                continue
            sheet[xlref(i + 1, 3)] = artist
            sheet[xlref(i + 1, 4)] = title

        self.create_episode_tabs()
        self.excel_save()

    def create_episode_tabs(self) -> None:
        """
        Per Episode a report is produced which shows how long clips have been used
        and which artist and title belongs to the clips.
        """
        duration_data = self.get_duration_data_per_episode_per_clip()

        # Delete old tabs
        for sheet_name in self.wb.sheetnames:
            if sheet_name == self.source_sheet_name or sheet_name == self.GENERAL_ARTIST_TITLE_OVERVIEW_NAME \
                    or sheet_name == self.DB_SHEET_NAME:
                continue
            self.wb.remove(self.wb[sheet_name])

        cache = self.get_artist_title_cache_from_clip_name()
        for episode_name in duration_data:

            sheet = self.wb.create_sheet(episode_name)
            index = 0
            sheet[xlref(0, 0)] = 'Clip name'
            sheet[xlref(0, 1)] = 'Artist'
            sheet[xlref(0, 2)] = 'Title'
            sheet[xlref(0, 3)] = 'Duration'

            for clip_name in duration_data[episode_name]:

                artist, title = cache[clip_name]

                sheet[xlref(index + 1, 0)] = clip_name
                sheet[xlref(index + 1, 1)] = artist
                sheet[xlref(index + 1, 2)] = title
                sheet[xlref(index + 1, 3)] = duration_data[episode_name][clip_name]
                index += 1
            custom_layout_sheet(sheet)
        self.excel_save()

    def create_or_update_db_sheet(self, update: bool = False) -> None:
        """
        Creates or updates dB sheet given the original data sheet. This sheet is used as
        dB table for the App.
        """
        db_data = self.read_db_sheet() if update is True else []
        source_data = self.read_source_sheet()

        if update is False:
            if self.DB_SHEET_NAME in self.wb.sheetnames:
                raise SheetExistsException("Database Sheet exists. Remove manually before new version can be created")

        for sheet_name in self.wb.sheetnames:
            if sheet_name == self.source_sheet_name:
                continue
            self.wb.remove(self.wb[sheet_name])

        sheet = self.wb.create_sheet(self.DB_SHEET_NAME)

        sheet[xlref(0, 0)] = 'Episode name'
        sheet[xlref(0, 1)] = 'Event'
        sheet[xlref(0, 2)] = 'Clip name'
        sheet[xlref(0, 3)] = 'Artist'
        sheet[xlref(0, 4)] = 'Title'
        sheet[xlref(0, 5)] = 'Start time'
        sheet[xlref(0, 6)] = 'End time'
        sheet[xlref(0, 7)] = 'Duration'

        cache = self.get_artist_title_cache_from_clip_name()
        for index, (source_row, db_row) in enumerate(zip_longest(source_data, db_data)):
            clip_name = self.get_clip_name_from_raw_clip_name(source_row.raw_clip_name)
            episode_name = self.get_episode_name(source_row.episode_index)
            sheet[xlref(index + 1, 0)] = episode_name
            sheet[xlref(index + 1, 1)] = source_row.event
            sheet[xlref(index + 1, 2)] = clip_name
            sheet[xlref(index + 1, 5)] = ':'.join(source_row.start.split(':')[1::])
            sheet[xlref(index + 1, 6)] = ':'.join(source_row.end.split(':')[1::])
            sheet[xlref(index + 1, 7)] = source_row.duration
            if update is False:
                sheet[xlref(index + 1, 3)] = None
                sheet[xlref(index + 1, 4)] = None
            else:
                if db_row is None:
                    artist, title = cache[clip_name]
                    sheet[xlref(index + 1, 3)] = artist
                    sheet[xlref(index + 1, 4)] = title
                else:
                    sheet[xlref(index + 1, 3)] = db_row.artist
                    sheet[xlref(index + 1, 4)] = db_row.title

        custom_layout_sheet(sheet)
        self.excel_save()

    def create_clip_usage_per_episode_sheet(self) -> None:
        """
        Generates a sheet which provides information what clip is used in what Episode
        """
        clip_names = set()
        clip_name_in_episodes = defaultdict(set)
        episode_indices = set()

        for row in self.read_source_sheet():
            clip_name = self.get_clip_name_from_raw_clip_name(row.raw_clip_name)
            clip_names.add(clip_name)
            clip_name_in_episodes[clip_name].add(self.get_episode_name(row.episode_index))
            episode_indices.add(self.get_episode_name(row.episode_index))

        for sheet_name in self.wb.sheetnames:
            if sheet_name == self.source_sheet_name or sheet_name == self.DB_SHEET_NAME:
                continue
            self.wb.remove(self.wb[sheet_name])

        sheet = self.wb.create_sheet(self.GENERAL_ARTIST_TITLE_OVERVIEW_NAME)

        sheet[xlref(0, 0)] = 'Clip name'

        for index, episode_index in enumerate(sorted(list(episode_indices))):
            sheet[xlref(0, 1 + index)] = episode_index

        for index, clip_name in enumerate(sorted(list(clip_names))):
            sheet[xlref(index + 1, 0)] = clip_name
            for j, episode_index in enumerate(sorted(list(episode_indices))):
                sheet[xlref(index + 1, 1 + j)] = 'x' if episode_index in clip_name_in_episodes[clip_name] else None

        custom_layout_sheet(sheet)
        self.excel_save()

    def split_mp4_file(self, episode_name: str) -> None:
        """
        Given the original MP4 file of an episode, this file can be split into
        separate files given the start and end time of a given event.
        """
        try:
            from pydub import AudioSegment

            event_data = {}
            for row in self.read_source_sheet():
                if row.episode_name != episode_name or row.event is None:
                    continue
                event_data[row.event] = row.start, row.end

            os.chdir(self.project_root_dir + '/' + episode_name)

            # Create object to create audio fragments
            audio_source = AudioSegment.from_file("Episode.mp4", "mp4")

            for event in event_data:

                raw_start_time = event_data[event][0]
                raw_end_time = event_data[event][1]

                # strip irrelevant episode index from time
                raw_start_time = ':'.join(raw_start_time.split(':')[1::])
                raw_end_time = ':'.join(raw_end_time.split(':')[1::])

                start_time = self.convert_time_to_ms(raw_start_time)
                end_time = self.convert_time_to_ms(raw_end_time)

                cur_song = audio_source[start_time:end_time]
                cur_song.export(f'Event{event}.mp4', format='mp4')

        except Exception:
            raise  # Raise, as this is run in the GUI in a custom thread where the exception is handled.
        finally:
            os.chdir(self.ROOT_DIR)

    def get_clip_name_from_raw_clip_name(self, raw_clip_name: str) -> str:
        """
        Helper function which strips the file format (.mp3, mp4, etc.) from a "raw" clip name.
        """
        clip_name = None
        found = False
        for file_type in self.FILE_TYPES:
            m = re.match(rf'.+?(?=.{file_type})', raw_clip_name)
            if m:
                clip_name = m.group(0)
                found = True
        if not found and '.' in raw_clip_name or clip_name is None:
            return f"Could not determine clip name from {raw_clip_name}."
        else:
            return clip_name

    def get_clip_name_from_event(self, event: str) -> str:
        """
        Returns clip name from given event.
        """
        for row in self.read_db_sheet():
            if row.event == event:
                return row.clip_name

    @staticmethod
    def convert_time_to_ms(time: str) -> int:
        """
        Returns time in milliseconds given time string
        """
        minutes, seconds, frames = time.split(':')
        return round(1000*(int(frames)/25) + 1000*int(seconds) + 60000*int(minutes))

    def get_episode_names(self) -> list[str]:
        """
        Returns list of Episode names
        """
        return sorted(list({row.episode_name for row in self.read_db_sheet()}))

    def get_events_per_episode_name(self, episode_name: str) -> list[str]:
        """
        Returns list of Events given an Episode name
        """
        return [row.event for row in self.read_db_sheet() if row.episode_name == episode_name]

    @staticmethod
    def get_episode_name(episode_index: str) -> str:
        """
        Returns Episode name (eg. E01, E02, etc.) given the Episode index
        """
        return 'E' + episode_index if int(episode_index) > 9 else 'E0' + episode_index

    def get_artist_title_cache_from_clip_name(self) -> dict[str, tuple[str, str]]:
        """
        Returns artist and title information per clip name if present in dB sheet
        """
        cache = {}

        for row in self.read_db_sheet():
            cache[row.clip_name] = row.artist, row.title
        return cache

    def open_excel_document(self, excel_filename: str) -> None:
        """
        Opens Excel document given the absolute filename
        """
        self.wb = openpyxl.load_workbook(excel_filename)
        self.source_sheet_name = self.wb.sheetnames[0]

    def excel_save(self) -> None:
        """
        Saves Excel file
        """
        self.wb.save(self.excel_filename)

    # @staticmethod
    # def log_error(exception: Exception) -> None:
    #     method_name = inspect.currentframe().f_back.f_code.co_name
    #     logger.error(f"An error occurred in method {method_name}: {exception}")


def main():
    """
    Used for testing
    """
    pass
    # music_cue = MusicCue()


if __name__ == "__main__":
    main()
