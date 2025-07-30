import threading
from dataclasses import make_dataclass, dataclass

from openpyxl.utils import get_column_letter
from openpyxl import Workbook, worksheet
from openpyxl.styles import Font


def xlref(row: int, column: int, zero_indexed: bool = True) -> str:
    """
    openpyxl helper to generate Excel cell references.

    Args:
        row (int): Row index (zero-indexed or one-indexed based on zero_indexed).
        column (int): Column index (zero-indexed or one-indexed based on zero_indexed).
        zero_indexed (bool, optional): Whether the row and column indices are zero-indexed.
                                      Defaults to True.

    Returns:
        str: Excel cell reference (e.g., "A1", "B2").
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def custom_layout_sheet(sheet: worksheet) -> None:
    """
    Openpyxl helper to apply a custom layout to a worksheet.

    This function:
        - Freezes the first row.
        - Adds a filter to the entire sheet.
        - Auto-sizes columns based on cell content.
        - Makes the first row bold.
    """
    for i in range(0, sheet.max_column + 1):
        sheet.freeze_panes = xlref(1, i)

    sheet.auto_filter.ref = sheet.dimensions

    for letter in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(letter)
        max_width = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = (max_width + 1) * 1.50

    for cell in sheet[1]:  # Make first row bold
        cell.font = Font(bold=True)


def read_excel_tab(wb: Workbook, sheet_name: str, fields: list[tuple[str, str]]) -> list[dataclass]:

    """
    Reads data from an Excel sheet and returns a list of data class objects.
    All cells are read as string.

    Args:
        wb (openpyxl.Workbook): The Excel workbook object.
        sheet_name (str): The name of the sheet to read.
        fields (list): A list of tuples, where the first element of each tuple is the
                      column name in the sheet and the second element is the
                      corresponding attribute name for the data class.

    Returns:
        list: A list of data class objects containing the extracted data.
    Raises:
        Exception: If any type of error occurs during Excel data reading.
    """
    try:
        sheet = wb[sheet_name]

        col_name_to_col_index = {}
        for index, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
            if column[0].value:
                col_name_to_col_index[column[0].value.strip()] = index

        header_names = [element[0] for element in fields]
        attr_names = [element[1] for element in fields]

        data_class = make_dataclass('DataClass', attr_names)
        data = []
        for row in sheet.iter_rows(min_row=2):  # Skip the header row
            table_row = [str(cell.value).strip() if cell.value is not None else None
                         for cell in row]
            row_data = [table_row[col_name_to_col_index[header_name]] for header_name in header_names]
            if row_data:  # Skip empty rows
                data.append(data_class(*row_data))
        return data
    except Exception:
        raise


def read_original_excel_tab(wb: Workbook, sheet_name: str) -> list[dataclass]:

    """
    Reads data from an Excel sheet and returns a list of data class objects.
    All cells are read as string.

    Args:
        wb (openpyxl.Workbook): The Excel workbook object.
        sheet_name (str): The name of the sheet to read.

    Returns:
        list: A list of data class objects containing the extracted data.
    Raises:
        Exception: If any type of error occurs during Excel data reading.
    """
    try:
        sheet = wb[sheet_name]

        header_names = ['EVENT', 'CLIP NAME', 'START TIME', 'END TIME', 'DURATION', 'STATE']
        attr_names = ['episode_index', 'event', 'raw_clip_name', 'start', 'end', 'duration', 'state']

        col_name_to_col_index = {
            'CHANNEL': 0, 'EVENT': 1, 'CLIP NAME': 2, 'START TIME': 3, 'END TIME': 4, 'DURATION': 5, 'STATE': 6
        }
        data_class = make_dataclass('DataClass', attr_names)
        data = []
        epi_index = 0
        read = False
        for row in sheet.iter_rows():
            if row[0].value is None:
                continue
            if 'CHANNEL' in str(row[0].value):
                epi_index += 1
                read = True
                continue
            if isinstance(row[0].value, int) and read is True:
                table_row = [str(cell.value).strip() if cell.value is not None else None for cell in row]
                row_data = [table_row[col_name_to_col_index[header_name]] for header_name in header_names]
                row_data.insert(0, str(epi_index))
                data.append(data_class(*row_data))
            else:
                read = False
        return data
    except Exception:
        raise


class PropagateExceptionThread(threading.Thread):

    """
    Helper to propagate exceptions from within a thread to the caller environment using a queue
    """
    def __init__(self, group=None, target=None, args=(), kwargs=None, *, daemon=None, exception_queue=None):
        if kwargs is None:
            kwargs = {}
        super().__init__(group=group, target=target, args=args, kwargs=kwargs, daemon=daemon)
        self.exception_queue = exception_queue

    def run(self):
        try:
            super().run()
        except Exception as e:
            if self.exception_queue:
                self.exception_queue.put(e)
