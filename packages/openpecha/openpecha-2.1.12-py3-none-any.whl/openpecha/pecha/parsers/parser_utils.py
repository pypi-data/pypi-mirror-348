from pathlib import Path
from typing import Dict

import openpyxl

from openpecha.config import get_logger

logger = get_logger(__name__)


def extract_metadata_from_xlsx(input: Path):
    workbook = openpyxl.load_workbook(input)
    sheet = workbook.active

    metadata = {}
    # Iterate through rows, now including ZH column
    for row in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True
    ):
        key, bo_value, en_value, zh_value = row

        # Ensure key exists before adding to metadata
        if key:
            entry = {}
            if bo_value:
                entry["BO"] = bo_value.strip()
            if en_value:
                entry["EN"] = en_value.strip()
            if zh_value:
                entry["ZH"] = zh_value.strip()

            if entry:
                metadata[key] = entry

    language: Dict = metadata.get("language", {})
    input_lang = next(value for value in language.values() if value)
    metadata["language"] = input_lang

    metadata["title"] = metadata["title_short"]

    if "source" in metadata and isinstance(metadata["source"], dict):
        metadata["source"] = list(metadata["source"].values())[0]

    if "translation_of" in metadata and isinstance(metadata["translation_of"], dict):
        metadata["translation_of"] = list(metadata["translation_of"].values())[0]

    if "commentary_of" in metadata and isinstance(metadata["commentary_of"], dict):
        metadata["commentary_of"] = list(metadata["commentary_of"].values())[0]

    return metadata
