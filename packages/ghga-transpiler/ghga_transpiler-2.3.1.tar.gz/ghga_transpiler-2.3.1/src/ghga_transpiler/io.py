# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
#

"""IO related functionality"""

from __future__ import annotations

import json
import sys
from importlib import resources
from pathlib import Path
from typing import TextIO

from openpyxl import load_workbook

from .core import GHGAWorkbook


def read_workbook(
    path: Path, configs_package: resources.Package = "ghga_transpiler.configs"
) -> GHGAWorkbook:
    """Function to read-in a workbook"""
    return GHGAWorkbook(load_workbook(path), configs_package=configs_package)


def _write_json(data: dict, file: TextIO):
    """Write the data to the specified file in JSON format"""
    json.dump(obj=data, fp=file, ensure_ascii=False, indent=4)


def write_json(data: dict, path: Path | None, force: bool) -> None:
    """Write the data provided as a dictionary to the specified output path or
    to stdout if the path is None.
    """
    if path is None:
        _write_json(data, sys.stdout)
    elif path.exists() and not force:
        raise FileExistsError(f"File already exists: {path}")
    else:
        with open(file=path, mode="w", encoding="utf8") as outfile:
            _write_json(data, outfile)
