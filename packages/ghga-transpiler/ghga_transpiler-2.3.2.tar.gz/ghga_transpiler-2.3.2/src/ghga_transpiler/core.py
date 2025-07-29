# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
#

"""This module contains functionalities for processing excel sheets into json object."""

from __future__ import annotations

from collections.abc import Callable
from importlib import resources

import semver
from openpyxl import Workbook

from . import config


class InvalidSematicVersion(Exception):
    """Raised when a version string is invalid."""


class GHGAWorkbook:
    """A GHGA metadata XLSX workbook"""

    def __init__(self, workbook: Workbook, configs_package: resources.Package):
        """Create a new GHGAWorkbook object from an XLSX workbook"""
        self.workbook = workbook
        self.wb_version = GHGAWorkbook._get_version(workbook)
        self.config = config.load_config(self.major_minor_version, configs_package)

    @staticmethod
    def _get_version(workbook):
        """Function to get workbook version from the worksheet _properties"""
        if "__properties" in workbook.sheetnames:
            try:
                return semver.Version.parse(workbook["__properties"].cell(1, 1).value)
            except ValueError:
                raise InvalidSematicVersion(
                    "Unable to extract metadata model version from the provided workbook (not a valid semantic version)."
                ) from None
        raise SyntaxError(
            "Unable to extract metadata model version from the provided workbook (missing)."
        )

    @property
    def major_minor_version(self):
        """Returns only major and minor version numbers"""
        return f"{self.wb_version.major}.{self.wb_version.minor}"


def get_worksheet_rows(
    worksheet,
    min_row: int | None,
    max_row: int,
    min_col: int | None,
    max_col: int | None,
) -> list:
    """Function to create a list of rows of a worksheet"""
    return list(
        row
        for row in worksheet.iter_rows(
            min_row, max_row, min_col, max_col, values_only=True
        )
        if not all(cell is None for cell in row)
    )


def get_header(
    worksheet,
    header_row: int | None,
    min_col: int | None,
    max_col: int | None,
) -> list[str]:
    """Function to return a list column names of a worksheet"""
    return list(
        cell.value
        for row in worksheet.iter_rows(header_row, header_row, min_col, max_col)
        for cell in row
    )


def convert_rows(header, rows) -> list[dict]:
    """Function to return list of dictionaries, rows as worksheet row values and
    column names as keys
    """
    return [
        {
            key: value
            for key, value in zip(header, row)
            if value is not None and value != ""
        }
        for row in rows
    ]


def transform_rows(
    rows: list[dict], transformations: dict[str, Callable] | None
) -> list[dict]:
    """Transforms row values if it is applicable with a given function"""
    transformed = []
    for row in rows:
        transformed_row = {}
        for key, value in row.items():
            if transformations and key in transformations:
                transformed_row[key] = transformations[key](value)
            else:
                transformed_row[key] = value
        transformed.append(transformed_row)
    return transformed


def convert_workbook(ghga_workbook: GHGAWorkbook) -> dict:
    """Function to convert an input spreadsheet into JSON"""
    converted_workbook = {}
    for sheet in ghga_workbook.config.worksheets:
        if sheet.settings is not None:
            if sheet.sheet_name in ghga_workbook.workbook:
                rows = get_worksheet_rows(
                    ghga_workbook.workbook[sheet.sheet_name],
                    sheet.settings.start_row,
                    ghga_workbook.workbook[sheet.sheet_name].max_row,
                    sheet.settings.start_column,
                    sheet.settings.end_column,
                )

                header = get_header(
                    ghga_workbook.workbook[sheet.sheet_name],
                    sheet.settings.header_row,
                    sheet.settings.start_column,
                    sheet.settings.end_column,
                )
                converted_rows = convert_rows(header, rows)
                transformed_rows = transform_rows(
                    converted_rows, sheet.settings.transformations
                )
                converted_workbook[sheet.settings.name] = transformed_rows
            else:
                converted_workbook[sheet.settings.name] = []

        else:
            raise ValueError(f"{sheet.settings} will never be None")
    return converted_workbook
